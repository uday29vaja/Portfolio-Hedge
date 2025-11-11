# web_app_local_dynamic.py - Portfolio Beta & Hedging with Custom Volatility & Risk-Free Rate
import json
import math
import os
import time
import requests
import streamlit as st
import pandas as pd
import numpy as np
import yfinance as yf
from datetime import datetime, timezone, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font
from scipy.stats import norm
from math import log, sqrt, exp
# -------------------------------
# Config
# -------------------------------
END_DATE = datetime.now(timezone.utc).date()
START_DATE = END_DATE - timedelta(days=365)
YAHOO_INDEX_TICKER = "^NSEI"

# -------------------------------
# Helper Functions
# -------------------------------
def download_yahoo_adjclose(ticker, start, end):
    try:
        data = yf.download(
            ticker,
            start=start.isoformat(),
            end=(end + timedelta(days=1)).isoformat(),
            progress=False,
            threads=False,
            auto_adjust=True
        )
        if data is None or data.empty:
            return 0
        return data.get("Adj Close") or data.get("Close")
    except Exception as e:
        st.warning(f"Failed to fetch {ticker}: {e}")
        return 0

def compute_beta(stock_series, index_series):
    df = pd.concat([stock_series, index_series], axis=1, join="inner").dropna()
    df = df.sort_index()
    if df.shape[0] < 30:
        return np.nan
    returns = df.pct_change().dropna()
    if returns.shape[0] < 20:
        return np.nan
    cov = returns.cov().iloc[0,1]
    var_index = returns.iloc[:,1].var()
    return cov/var_index if var_index != 0 else np.nan

def get_stock_beta(symbol, index_series):
    yf_ticker = f"{symbol}.NS"
    series = download_yahoo_adjclose(yf_ticker, START_DATE, END_DATE)
    if series is None or series.empty:
        return symbol, np.nan
    beta = compute_beta(series, index_series)
    return symbol, beta

def black_scholes_put_price(S, K, T, r, sigma):
    d1 = (log(S/K) + (r + 0.5*sigma**2)*T) / (sigma * sqrt(T))
    d2 = d1 - sigma * sqrt(T)
    return K * exp(-r * T) * norm.cdf(-d2) - S * norm.cdf(-d1)


# 🔹 Common utility
def get_last_tuesday(year, month):
    """Return the last Tuesday of a given month."""
    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)

    last_day = next_month - timedelta(days=1)
    while last_day.weekday() != 1:  # Tuesday = 1
        last_day -= timedelta(days=1)
    return last_day
# 🔹 Monthly expiry
def get_monthly_expiry(df, symbol="NIFTY"):
    """Get last Tuesday of current month (Monthly expiry).
       If expiry is today or within 1 day, return next month's expiry."""
    if df is None or df.empty:
        print("⚠️ No option chain data available.")
        return None

    # Convert expiry strings to datetime (keep as datetime, not date)
    df['expiry'] = pd.to_datetime(df['expiry'], format='%d-%b-%Y', errors='coerce')
    if df['expiry'].isna().all():
        print("⚠️ No valid expiry dates found.")
        return None

    today = datetime.now()
    curr_year, curr_month = today.year, today.month

    # Current month expiries
    curr_expiries = sorted(
        e for e in df['expiry']
        if e.year == curr_year and e.month == curr_month
    )

    if not curr_expiries:
        print("⚠️ No expiries found for current month.")
        return None

    monthly_expiry = curr_expiries[-1]

    # Roll to next month if <15 days remain
    if (monthly_expiry - today).days <= 15:
        next_month = (curr_month % 12) + 1
        next_year = curr_year + (1 if next_month == 1 else 0)
        next_expiries = sorted(
            e for e in df['expiry']
            if e.year == next_year and e.month == next_month
        )
        if next_expiries:
            monthly_expiry = next_expiries[-1]

    print(f"✅ Monthly expiry for {symbol}: {monthly_expiry.strftime('%d-%b-%Y')}")
    return monthly_expiry
 

# 🔹 Quarterly expiry
def get_quarterly_expiry(df, symbol="NIFTY"):
    if df is None or df.empty:
        print("⚠️ No option chain data available.")
        return None

    # Convert expiry strings to datetime objects
    df['expiry'] = pd.to_datetime(df['expiry'], format='%d-%b-%Y', errors='coerce')
    if df['expiry'].isna().all():
        print("⚠️ No valid expiry dates found.")
        return None

    today = datetime.now()
    curr_year, curr_month = today.year, today.month

    # Determine current quarter months
    if curr_month <= 3:
        quarter_months = [1,2,3]
    elif curr_month <= 6:
        quarter_months = [4,5,6]
    elif curr_month <= 9:
        quarter_months = [7,8,9]
    else:
        quarter_months = [10,11,12]

    # Filter only expiries in current quarter and >= today
    quarter_expiries = df[
        (df['expiry'].dt.year == curr_year) &
        (df['expiry'].dt.month.isin(quarter_months)) &
        (df['expiry'].dt.date >= today.date())
    ]['expiry'].unique()

    if len(quarter_expiries) == 0:
        print("⚠️ No expiries found for current quarter.")
        return None

    # Last expiry in the quarter is the quarterly expiry
    quarterly_expiry = max(quarter_expiries)

    print(f"✅ Current quarter ({curr_year} Q) expiry for {symbol}: {quarterly_expiry.strftime('%d-%b-%Y')}")
    return quarterly_expiry


# 🔹 Annual expiry
def get_annual_expiry(df, symbol="NIFTY"):
    if df is None or df.empty:
        print("⚠️ No option chain data available.")
        return None

    # Convert expiry strings to datetime
    df['expiry'] = pd.to_datetime(df['expiry'], format='%d-%b-%Y', errors='coerce')
    if df['expiry'].isna().all():
        print("⚠️ No valid expiry dates found.")
        return None

    today = datetime.now()
    year = today.year

    # Filter for December expiries of current year >= today
    dec_expiries = df[
        (df['expiry'].dt.year == year) &
        (df['expiry'].dt.month == 12) &
        (df['expiry'].dt.date >= today.date())
    ]['expiry'].unique()

    # If no remaining December expiry, look at next year's December
    if len(dec_expiries) == 0:
        year += 1
        dec_expiries = df[
            (df['expiry'].dt.year == year) &
            (df['expiry'].dt.month == 12)
        ]['expiry'].unique()

    if len(dec_expiries) == 0:
        print("⚠️ No December expiry found.")
        return None

    # Annual expiry = last expiry in December
    annual_expiry = max(dec_expiries)

    print(f"✅ Annual expiry for {symbol} ({year} Dec): {annual_expiry.strftime('%d-%b-%Y')}")
    return annual_expiry

def get_implied_volatility(S, K, T, r, market_price, tol=1e-6, max_iter=100):
    print("Calculating Implied Volatility...", S, K, T, r, market_price)
    sigma = 0.2  # initial guess
    for i in range(max_iter):
        price = black_scholes_put_price(S, K, T, r, sigma)
        d1 = (log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * sqrt(T))
        vega = S * norm.pdf(d1) * sqrt(T)
        diff = market_price - price
        if abs(diff) < tol:
            return sigma
        sigma += diff / vega
        sigma = max(sigma, 0.001)
    return sigma
# --- Step 2: Fetch NSE NIFTY put premium ---
import requests
import json
import time
import nsefin

def fetch_option_chain(symbol: str):
    """Fetch full option chain data from NSE using nsefin."""
    try:
        print(f"📡 Fetching option chain for {symbol} via nsefin...")
        nse = nsefin.NSEClient()
        option_chain = nse.get_option_chain(symbol)
        print("✅ Option chain fetched successfully.")
        return option_chain  # DataFrame
    except Exception as e:
        print("❌ Error fetching option chain:", e)
        return None
    
def fetch_put_premium(df, strike_price, expiry_date):
    """
    Extract PUT premium (pe_ltp) for the given strike and expiry.
    expiry_date format: '25-Nov-2025'
    """
    if df is None or df.empty:
        print("⚠️ Empty option chain data.")
        return None

    expiry_str = expiry_date.strip().upper()
    strike_price = float(strike_price)

    # Normalize column names
    df.columns = [c.strip().lower() for c in df.columns]

    required_cols = {'strike', 'expiry', 'pe_ltp'}
    if not required_cols.issubset(df.columns):
        print("⚠️ Missing expected columns:", df.columns)
        return None

    # Convert expiry for matching
    df['expiry'] = df['expiry'].astype(str).str.upper()

    # Filter for matching strike and expiry
    df_filtered = df[
        (df['strike'] == strike_price) &
        (df['expiry'] == expiry_str)
    ]

    if not df_filtered.empty:
        last_price = df_filtered.iloc[0]['pe_ltp']
        print(f"✅ Found PUT {strike_price} @ {expiry_str}: {last_price}")
        return float(last_price)

    print(f"⚠️ No PUT found for strike {strike_price} @ {expiry_str}")
    return None
# -------------------------------
# Hedging Calculation

def calculate_hedging(portfolio_beta, total_value, hedge_percentage):
    print("Calculating Hedging Costs...")
    r = 0.06  # Risk-free rate
    nifty = yf.Ticker("^NSEI")   # Symbol for NIFTY 50
    nifty_price = round(nifty.history(period="1d")["Close"].iloc[-1],2)
    print("NIFTY Price:", nifty_price)

    today = datetime.today()

    expiry_data = fetch_option_chain("NIFTY")
    if expiry_data is None:
        print("❌ Failed to fetch option chain data.")
        return None
    
    hedge_exposure = total_value * portfolio_beta * (hedge_percentage / 100)
    monthly_expiry_date = get_monthly_expiry(expiry_data)
    quarterly_expiry_date = get_quarterly_expiry(expiry_data)
    annual_expiry_date = get_annual_expiry(expiry_data)
    #  Monthly strike: always in 100s
    monthly_strike = math.floor(nifty_price / 100) * 100
    print("monthly_expiry_date:", monthly_expiry_date)
    
    
    # Quarterly strike: based on expiry month``
    if quarterly_expiry_date.month == 12:
        quarterly_strike = math.floor(monthly_strike / 1000) * 1000

    else:
        quarterly_strike = math.floor(monthly_strike / 100) * 100

    # Annual strike: based on expiry month
    if annual_expiry_date.month == 12:
        annual_strike = math.floor(monthly_strike / 1000) * 1000

    else:
        annual_strike = math.floor(monthly_strike / 100) * 100

    monthly_T = round( (monthly_expiry_date - today).days / 365 ,2)
    quarterly_T = round( (quarterly_expiry_date  - today).days / 365 ,2)    
    annual_T = round( (annual_expiry_date  - today).days / 365 ,2)

    expiry_data = fetch_option_chain("NIFTY")
    if expiry_data is None:
        print("❌ Failed to fetch option chain data.")
        return None
    Monthly_put_premium = fetch_put_premium(expiry_data, monthly_strike , monthly_expiry_date.strftime("%d-%b-%Y"))
    quarterly_put_premium = fetch_put_premium(expiry_data, quarterly_strike , quarterly_expiry_date.strftime("%d-%b-%Y"))
    annual_put_premium = fetch_put_premium(expiry_data, annual_strike , annual_expiry_date.strftime("%d-%b-%Y"))
    print("Monthly Put Premium:", Monthly_put_premium)
    print("Quarterly Put Premium:", quarterly_put_premium)
    print("Annual Put Premium:", annual_put_premium)    
   

    monthly_sigma = get_implied_volatility(nifty_price,monthly_strike,monthly_T,r,Monthly_put_premium)  # Assumed volatility
    quarterly_sigma = get_implied_volatility(nifty_price,quarterly_strike,quarterly_T,r,quarterly_put_premium)  # Assumed volatility
    annual_sigma = get_implied_volatility(nifty_price,annual_strike,annual_T,r,annual_put_premium)  # Assumed volatility
    print("Monthly Implied Volatility:", monthly_sigma)
    print("Quarterly Implied Volatility:", quarterly_sigma)
    print("Annual Implied Volatility:", annual_sigma)

     # Safe defaults if None
    if monthly_sigma is None or not isinstance(monthly_sigma, (int, float)):
        monthly_sigma = 0.15
    if quarterly_sigma is None or not isinstance(quarterly_sigma, (int, float)):
        quarterly_sigma = 0.15
    if annual_sigma is None or not isinstance(annual_sigma, (int, float)):
        annual_sigma = 0.15
    monthly_lot = math.ceil( hedge_exposure / (monthly_strike * 75))  # NIFTY lot size = 75
    quarterly_lot = math.ceil(hedge_exposure / (quarterly_strike * 75) )       
    annual_lot = math.ceil( hedge_exposure / (annual_strike * 75))
    print("Monthly Lots:", monthly_lot)
    print("Quarterly Lots:", quarterly_lot)
    print("Annual Lots:", annual_lot)

    print("----------------------------------------------------------------------------")
    monthly_cost = round( Monthly_put_premium * 75 * monthly_lot ,2) # NIFTY lot size = 75
    quarterly_cost = round( quarterly_put_premium * 75 * quarterly_lot ,2)
    annual_cost = round( annual_put_premium * 75 * annual_lot ,2)
    print("Monthly Cost:", monthly_cost)
    print("Quarterly Cost:", quarterly_cost)    
    print("Annual Cost:", annual_cost)


    print("----------------------------------------------------------------------------")
    Monthly_Annualised_premium = black_scholes_put_price(nifty_price, monthly_strike, 1, r, monthly_sigma)
    Quarterly_Annualised_premium = black_scholes_put_price(nifty_price, quarterly_strike, 1, r, quarterly_sigma)
    Annual_Annualised_premium = black_scholes_put_price(nifty_price, annual_strike, 1, r, annual_sigma)
    print("Monthly Annualised Premium (%):", Monthly_Annualised_premium)
    print("Quarterly Annualised Premium (%):", Quarterly_Annualised_premium)    
    print("Annual Annualised Premium (%):", Annual_Annualised_premium)  

    monthly_annualized = ((Monthly_Annualised_premium - max(monthly_strike - nifty_price,0)) / monthly_strike) *100  #(monthly_cost / total_value) * (365 / (monthly_T * 365)) * 100
    quarterly_annualized = ((Quarterly_Annualised_premium - max(quarterly_strike - nifty_price,0)) / quarterly_strike) *100   #(quarterly_cost / total_value) * (365 / (quarterly_T * 365)) * 100
    annual_annualized =  ((Annual_Annualised_premium - max(annual_strike - nifty_price,0)) / annual_strike) *100   #(annual_cost / total_value) * (365 / (annual_T * 365)) * 100
    print("Monthly Annualized Cost (%):", monthly_annualized)
    print("Quarterly Annualized Cost (%):", quarterly_annualized)   
    print("Annual Annualized Cost (%):", annual_annualized)

    scenarios = [-0.2, -0.1, 0, 0.1, 0.2]
    scenario_analysis = []
    for pct in scenarios:
        new_portfolio = total_value * (1 + pct)
        monthly_payoff = max(0, monthly_strike - new_portfolio)
        quarterly_payoff = max(0, quarterly_strike - new_portfolio)
        annual_payoff = max(0, annual_strike - new_portfolio)
        scenario_analysis.append({
            "period": "Monthly",
            "scenario": f"{pct*100:+.0f}%",
            "end_portfolio": new_portfolio,
            "put_payoff": monthly_payoff,
            "net_value_hedge": new_portfolio + monthly_payoff - monthly_cost
        })
        scenario_analysis.append({
            "period": "Quarterly",
            "scenario": f"{pct*100:+.0f}%",
            "end_portfolio": new_portfolio,
            "put_payoff": quarterly_payoff,
            "net_value_hedge": new_portfolio + quarterly_payoff - quarterly_cost
        })
        scenario_analysis.append({
            "period": "Annual",
            "scenario": f"{pct*100:+.0f}%",
            "end_portfolio": new_portfolio,
            "put_payoff": annual_payoff,
            "net_value_hedge": new_portfolio + annual_payoff - annual_cost
        })

    return {
        "monthly_put_strike": monthly_strike,
        "quarterly_put_strike": quarterly_strike,
        "annual_put_strike": annual_strike,
        "monthly_expiry": monthly_expiry_date.strftime("%d-%b-%Y") ,#(datetime.now() + timedelta(days=30)).strftime("%d-%b-%Y"),
        "quarterly_expiry": quarterly_expiry_date.strftime("%d-%b-%Y"), # (datetime.now() + timedelta(days=90)).strftime("%d-%b-%Y"),
        "annual_expiry": annual_expiry_date.strftime("%d-%b-%Y"), #(datetime.now() + timedelta(days=365)).strftime("%d-%b-%Y"),
        "monthly_cost": monthly_cost,
        "quarterly_cost": quarterly_cost, 
        "annual_cost": annual_cost,
        "monthly_annualized_cost": round(monthly_annualized,2),
        "quarterly_annualized_cost": round(quarterly_annualized,2),
        "annual_annualized_cost": round(annual_annualized,2),
        "monthly_lots": monthly_lot,
        "quarterly_lots": quarterly_lot,    
        "annual_lots": annual_lot,
        "monthly_premium": round(Monthly_put_premium,2),
        "quarterly_premium": round(quarterly_put_premium,2), 
        "annual_premium": round(annual_put_premium,2),
        "scenario_analysis": scenario_analysis
    }

# Excel Export
def create_excel_export(portfolio_data, hedging_data, portfolio_beta, total_amount, hedge_percentage):
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Summary"
    ws['A1'] = "Portfolio Beta & Hedging Results"
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = "Total Portfolio Value"
    ws['B3'] = f"₹{total_amount:,.2f}"
    ws['A4'] = "Hedge Percentage"
    ws['B4'] = f"{hedge_percentage}%"
    ws['A5'] = "Portfolio Beta"
    ws['B5'] = f"{portfolio_beta:.4f}"
    ws['A6'] = "Hedge Exposure"
    ws['B6'] = f"₹{total_amount * portfolio_beta * (hedge_percentage/100):,.2f}"

    # Portfolio Breakdown
    ws['A8'] = "Portfolio Breakdown"
    ws['A8'].font = Font(bold=True)
    for col, header in enumerate(portfolio_data.columns, 1):
        ws.cell(row=9, column=col, value=header).font = Font(bold=True)
    for row, (_, data) in enumerate(portfolio_data.iterrows(), 10):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)

    # Hedging Costs
    ws2 = wb.create_sheet("Hedging Costs")
    ws2['A1'] = "Hedging Costs"
    ws2['A1'].font = Font(bold=True, size=14)
    headers = ["Option Type", "Put Strike", "Expiry", "Cost", "Annualized Cost %"]
    for col, h in enumerate(headers, 1):
        ws2.cell(row=3, column=col, value=h).font = Font(bold=True)
    for i, period in enumerate(["monthly", "quarterly", "annual"], 4):
        ws2.cell(row=i, column=1, value=period.capitalize())
        ws2.cell(row=i, column=2, value=f"₹{hedging_data[f'{period}_put_strike']}")
        ws2.cell(row=i, column=3, value=hedging_data[f'{period}_expiry'])
        ws2.cell(row=i, column=4, value=f"₹{hedging_data[f'{period}_cost']:,.2f}")
        ws2.cell(row=i, column=5, value=f"{hedging_data[f'{period}_annualized_cost']:.2f}%")

    return wb

# -------------------------------
# Streamlit App
# -------------------------------
st.set_page_config(page_title="Portfolio Beta Calculator", layout="wide")
st.title("📊 Portfolio Beta & Hedging Calculator (Local, Custom Volatility)")

# Portfolio Input
st.header("1. Portfolio Input")
input_method = st.radio("Choose input method:", ["Manual Entry", "CSV Upload"])
portfolio_data = None
nse_symbols = pd.read_csv("EQUITY_L.csv")["SYMBOL"].tolist()

if input_method == "Manual Entry":
    
    st.subheader("Enter Stocks Manually")
    num_stocks = st.number_input("Number of stocks:", min_value=1, max_value=20, value=3)
    stocks = []
    for i in range(num_stocks):
        col1, col2 = st.columns(2)
        with col1:
            symbol = st.selectbox(
                f"Select Stock Symbol {i+1}",
                options=nse_symbols,
                index=nse_symbols.index("RELIANCE") if "RELIANCE" in nse_symbols else 0,
                key=f"sym_{i}"
            )
        with col2:
            amount = st.number_input(
                f"Investment Amount (₹) {i+1}",
                min_value=0,
                value=10000,
                step=1000,
                key=f"amt_{i}"
            )
        stocks.append({"SYMBOL": symbol, "AMOUNT": amount})
    portfolio_data = pd.DataFrame(stocks)
    st.write("Your Portfolio:")
    st.dataframe(portfolio_data)

else:
    st.subheader("Upload Portfolio CSV")
    st.info("Your CSV should have columns: SYMBOL, AMOUNT")

    # Create a sample CSV file in memory
    sample_data = pd.DataFrame({
        "SYMBOL": ["RELIANCE", "INFY"],
        "AMOUNT": [10000, 15000]
    })
    csv_buffer = io.StringIO()
    sample_data.to_csv(csv_buffer, index=False)

    # Download button for sample CSV
    st.download_button(
        label="📥 Download Sample CSV",
        data=csv_buffer.getvalue(),
        file_name="sample_portfolio.csv",
        mime="text/csv"
    )

    # Upload section
    uploaded_file = st.file_uploader("Choose Portfolio File", type=['csv', 'xlsx'])

    if uploaded_file is not None:
        if uploaded_file.name.endswith('.csv'):
            portfolio_data = pd.read_csv(uploaded_file)
        else:
            portfolio_data = pd.read_excel(uploaded_file)
            
        st.write("Uploaded Portfolio:")
        st.dataframe(portfolio_data)



# Hedge Calculation
if portfolio_data is not None:
    st.header("2. Hedge Settings & Black-Scholes Parameters")
     # Add Hedge Percentage Selection
    st.subheader("🛡️ Hedge Protection Level")
    col1, col2 = st.columns(2)
    with col1:
        hedge_percentage = st.selectbox(
            "How much portfolio do you want to hedge?",
            [100, 75, 50, 25],
            index=0,
            help="Select the percentage of your portfolio exposure you want to protect"
        )
        st.write(f"**Selected: {hedge_percentage}% protection**")
    
    with col2:
        st.info(f"""
        **Hedge Percentage Guide:**
        - **100%**: Full protection (most expensive)
        - **75%**: Balanced protection
        - **50%**: Moderate protection  
        - **25%**: Basic protection (least expensive)
        """)

    if st.button("🚀 Calculate Beta & Hedging", type="primary"):
        if "AMOUNT" not in portfolio_data.columns:
            st.error("❌ Portfolio must have 'AMOUNT' column")
        else:
            try:
                # Convert AMOUNT to numbers
                portfolio_data["AMOUNT"] = pd.to_numeric(portfolio_data["AMOUNT"], errors='coerce')
                print("Portfolio Data after conversion:", portfolio_data)
                if portfolio_data["AMOUNT"].isna().any():
                    st.error("❌ Some AMOUNT values are not valid numbers.")
                else:
                    with st.spinner("📊 Calculating betas and advanced hedging costs..."):
                        # Calculate weights and beta
                        total_amount = portfolio_data["AMOUNT"].sum()
                        if total_amount == 0:
                            st.error("❌ Total portfolio amount cannot be zero")
                        else:
                            portfolio_data["WEIGHT"] = portfolio_data["AMOUNT"] / total_amount
                            print("Portfolio Data with Weights:", portfolio_data)   
                            # Download index data and calculate beta
                            index_series = download_yahoo_adjclose(YAHOO_INDEX_TICKER, START_DATE, END_DATE)
                            
                            if index_series is None or index_series.empty:
                                st.error("❌ Failed to download index data.")
                            else:
                                index_series = index_series.dropna().sort_index()

                                # Calculate betas
                                betas = []
                                for sym in portfolio_data["SYMBOL"]:
                                    symbol, beta = get_stock_beta(sym, index_series)
                                    betas.append((symbol, beta))
                                
                                # Merge results
                                beta_df = pd.DataFrame(betas, columns=["SYMBOL", "BETA"])
                                merged = pd.merge(portfolio_data, beta_df, on="SYMBOL", how="left")
                                merged["WEIGHTED_BETA"] = merged["WEIGHT"] * merged["BETA"]
                                portfolio_beta = merged["WEIGHTED_BETA"].sum()

                                # Local Hedging
                                hedging_data = calculate_hedging(portfolio_beta, total_amount, hedge_percentage)

                                # Display Results
                                st.header("3. Advanced Hedging Results")
                                st.success("✅ Calculation Complete!")
                                
                                # Protection Level Info
                                st.subheader("🛡️ Protection Details")
                                col1, col2, col3, col4 = st.columns(4)
                                with col1:
                                    st.metric("Total Portfolio Value", f"₹{total_amount:,.2f}")
                                with col2:
                                    st.metric("Hedge Percentage", f"{hedge_percentage}%")
                                with col3:
                                    st.metric("Portfolio Beta", f"{portfolio_beta:.4f}")
                                with col4:
                                    st.metric("Hedge Exposure", f"₹{total_amount * portfolio_beta * (hedge_percentage/100):,.2f}")
                                
                                # hedging cost metrics
                                st.subheader("💰 Hedging Costs & Details")

                                col1, col2, col3 = st.columns(3)

                                # 🗓️ Monthly
                                with col1:
                                    st.markdown("#### 🗓️ Monthly")
                                    st.metric("Put Strike (₹)", f"{hedging_data['monthly_put_strike']:,}")
                                    st.metric("Expiry", hedging_data['monthly_expiry'])
                                    st.metric("Cost", f"₹{hedging_data['monthly_cost']:,.2f}")
                                    st.metric("Annualized Cost %", f"{hedging_data['monthly_annualized_cost']:.2f}%")
                                    st.metric("Lots Required", hedging_data['monthly_lots'])
                                    st.metric("Premium per Lot", f"₹{hedging_data['monthly_premium']}")
                                
                                # 📅 Quarterly
                                with col2:
                                    st.markdown("#### 📅 Quarterly")
                                    st.metric("Put Strike (₹)", f"{hedging_data['quarterly_put_strike']:,}")
                                    st.metric("Expiry", hedging_data['quarterly_expiry'])
                                    st.metric("Cost", f"₹{hedging_data['quarterly_cost']:,.2f}")
                                    st.metric("Annualized Cost %", f"{hedging_data['quarterly_annualized_cost']:.2f}%")
                                    st.metric("Lots Required", hedging_data['quarterly_lots'])
                                    st.metric("Premium per Lot", f"₹{hedging_data['quarterly_premium']}")
                                                            
                                # 🗓️ Annual
                                with col3:
                                    st.markdown("#### 🗓️ Annual")
                                    st.metric("Put Strike (₹)", f"{hedging_data['annual_put_strike']:,}")
                                    st.metric("Expiry", hedging_data['annual_expiry'])
                                    st.metric("Cost", f"₹{hedging_data['annual_cost']:,.2f}")
                                    st.metric("Annualized Cost %", f"{hedging_data['annual_annualized_cost']:.2f}%")
                                    st.metric("Lots Required", hedging_data['annual_lots'])
                                    st.metric("Premium per Lot", f"₹{hedging_data['annual_premium']}")


                                # Portfolio Breakdown
                                st.subheader("📈 Portfolio Breakdown")
                                st.dataframe(merged)

                                # Scenario Analysis
                                if hedging_data['scenario_analysis']:
                                    st.subheader("🎯 Scenario Analysis")
                                    scenario_df = pd.DataFrame(hedging_data['scenario_analysis'])
                                    
                                    # Display by period
                                    for period in ['Monthly', 'Quarterly', 'Annual']:
                                        period_data = scenario_df[scenario_df['period'] == period]
                                        if not period_data.empty:
                                            st.write(f"**{period} Hedging Scenarios:**")
                                            display_data = period_data.drop('period', axis=1)
                                            st.dataframe(display_data, width='stretch')
                                else:
                                    st.info("📊 Scenario analysis data will be available when calculations are complete")
                                
                                # Download Options
                                st.subheader("📥 Download Results")
                                csv = portfolio_data.to_csv(index=False)
                                st.download_button("Download Portfolio CSV", csv, "portfolio_results.csv", "text/csv")

                                excel_wb = create_excel_export(portfolio_data, hedging_data, portfolio_beta, total_amount, hedge_percentage)
                                buffer = io.BytesIO()
                                excel_wb.save(buffer)
                                buffer.seek(0)
                                st.download_button("Export Full Excel Report", buffer, "portfolio_hedging.xlsx",
                                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"❌ Error during calculation: {e}")   



# # -------------------------------
# # Equity Symbols Reference
# st.header("📋 Equity Symbols Reference")

# try:
#     # Check if file exists and load it
#     if os.path.exists("EQUITY_L.csv"):
#         scrips_df = pd.read_csv("EQUITY_L.csv")
#         total_symbols = len(scrips_df)
        
#         st.success(f"**{total_symbols} equity symbols available**")
#         st.write("Reference list of all available trading symbols")
        
#         # Download option
#         with open("EQUITY_L.csv", "rb") as file:
#             file_bytes = file.read()
        
#         st.download_button(
#             label="📥 Download Equity Symbols (CSV)",
#             data=file_bytes,
#             file_name="EQUITY_Symbols.csv",
#             mime="text/csv"
#         )
#     else:
#         st.info("Equity symbols file not found - download available when file is present")

# except Exception as e:
#     st.info("Equity reference data will be available when the symbols file is present")

