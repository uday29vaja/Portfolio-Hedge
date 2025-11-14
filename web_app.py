# web_app_local_dynamic.py - Portfolio Beta & Hedging with Custom Volatility & Risk-Free Rate
import json
import logging
import math
import os
import time
import requests
import streamlit as st
import pandas as pd
import numpy as np
import urllib3
import yfinance as yf
from datetime import datetime, timezone, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font
from scipy.stats import norm
from math import log, sqrt, exp
from zeep import Client, Transport
# -------------------------------
# Config
# -------------------------------
END_DATE = datetime.now(timezone.utc).date()
START_DATE = END_DATE - timedelta(days=365)
YAHOO_INDEX_TICKER = "^NSEI"

# Create logs directory
if not os.path.exists("logs"):
    os.makedirs("logs")

# Log file name per day
log_filename = f"logs/app_{datetime.now().strftime('%Y_%m_%d')}.log"

logging.basicConfig(
    filename=log_filename,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    filemode="a"
)

logger = logging.getLogger()
# -------------------------------
# Helper Functions
# -------------------------------
def download_yahoo_adjclose(ticker, start, end):
    try:
        data = yf.download(
            ticker,
            start=start.isoformat(),
            end=(end + timedelta(days=1)).isoformat(),
            progress=False,
            threads=False,
            auto_adjust=True
        )
        if data is None or data.empty:
            return 0
        return data.get("Adj Close") or data.get("Close")
    except Exception as e:
        st.warning(f"Failed to fetch {ticker}: {e}")
        logger.error(f"Failed to fetch {ticker}: {e}")
        return 0

def compute_beta(stock_series, index_series):
    df = pd.concat([stock_series, index_series], axis=1, join="inner").dropna()
    df = df.sort_index()
    if df.shape[0] < 30:
        return np.nan
    returns = df.pct_change().dropna()
    if returns.shape[0] < 20:
        return np.nan
    cov = returns.cov().iloc[0,1]
    var_index = returns.iloc[:,1].var()
    return cov/var_index if var_index != 0 else np.nan

def get_stock_beta(symbol, index_series):
    yf_ticker = f"{symbol}.NS"
    series = download_yahoo_adjclose(yf_ticker, START_DATE, END_DATE)
    if series is None or series.empty:
        return symbol, np.nan
    beta = compute_beta(series, index_series)
    return symbol, beta

def get_nav_data( scheme_code):
        """Get historical NAV data for a scheme"""
        try:
            url = f"https://api.mfapi.in/mf/{scheme_code}"
            response = requests.get(url, timeout=30)
            if response.status_code == 200:
                data = response.json()
                if 'data' in data and data['data']:
                    df = pd.DataFrame(data['data'])
                    df['date'] = pd.to_datetime(df['date'], format='%d-%m-%Y')
                    df['nav'] = pd.to_numeric(df['nav'], errors='coerce')
                    df.set_index('date', inplace=True)
                    df = df.sort_index().dropna()
                    
                    # Filter for last 2 years
                    two_years_ago = datetime.now() - timedelta(days=730)
                    df = df[df.index >= two_years_ago]
                    
                    return df['nav']
            return None
        except Exception as e:
            print(f"❌ Error fetching NAV for {scheme_code}: {e}")
            logger.error(f"❌ Error fetching NAV for {scheme_code}: {e}")
            return None

# Excel Export
def create_excel_export(portfolio_data, hedging_data, portfolio_beta, total_amount, hedge_percentage):
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Summary"
    ws['A1'] = "Portfolio Beta & Hedging Results"
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = "Total Portfolio Value"
    ws['B3'] = f"₹{total_amount:,.2f}"
    ws['A4'] = "Hedge Percentage"
    ws['B4'] = f"{hedge_percentage}%"
    ws['A5'] = "Portfolio Beta"
    ws['B5'] = f"{portfolio_beta:.4f}"
    ws['A6'] = "Hedge Exposure"
    ws['B6'] = f"₹{total_amount * portfolio_beta * (hedge_percentage/100):,.2f}"

    # Portfolio Breakdown
    ws['A8'] = "Portfolio Breakdown"
    ws['A8'].font = Font(bold=True)
    for col, header in enumerate(portfolio_data.columns, 1):
        ws.cell(row=9, column=col, value=header).font = Font(bold=True)
    for row, (_, data) in enumerate(portfolio_data.iterrows(), 10):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)

    # Hedging Costs
    ws2 = wb.create_sheet("Hedging Costs")
    ws2['A1'] = "Hedging Costs"
    ws2['A1'].font = Font(bold=True, size=14)
    headers = ["Option Type", "Put Strike", "Expiry", "Cost", "Annualized Cost %"]
    for col, h in enumerate(headers, 1):
        ws2.cell(row=3, column=col, value=h).font = Font(bold=True)
    for i, period in enumerate(["monthly", "quarterly", "annual"], 4):
        ws2.cell(row=i, column=1, value=period.capitalize())
        ws2.cell(row=i, column=2, value=f"₹{hedging_data[f'{period}_put_strike']}")
        ws2.cell(row=i, column=3, value=hedging_data[f'{period}_expiry'])
        ws2.cell(row=i, column=4, value=f"₹{hedging_data[f'{period}_cost']:,.2f}")
        ws2.cell(row=i, column=5, value=f"{hedging_data[f'{period}_annualized_cost']:.2f}%")

    return wb

# Disable SSL warnings for localhost only
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

#fetch data from API
def Get_hedge_data(portfolio_beta, total_value, hedge_percentage):
    print("Fetching hedging data from local API...")

    # Disable SSL warnings (for self-signed localhost certs)
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    # Prepare a session that ignores SSL certificate errors
    session = requests.Session()
    session.verify = False

    # Pass it to Zeep transport
    transport = Transport(session=session)

    # Create SOAP client
    wsdl = "https://portfoliohedge.finideas.com/PortFolioPayout/PortfolioService.asmx?WSDL"
    client = Client(wsdl=wsdl, transport=transport)

    # Call your web method
    result = client.service.hedge_calculation(portfolio_beta, total_value, hedge_percentage)
    

    # Usually result is a JSON string — convert it
    try:
        data = json.loads(result)
    except Exception as e:
        print("⚠️ Failed to parse JSON:", e)
        print("Raw result:", result)
        logger.error("Failed to parse JSON:", e)
        return None

    print("✅ Hedging data received successfully.")
    logger.info("✅ Hedging data received successfully.")

    return data
# Mutual Fund Beta Mapping (Category-based fallback)

def Get_EQSymbol():
    print("Fetching EQ Symbol from local API...")

    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    session = requests.Session()
    session.verify = False
    transport = Transport(session=session)

    wsdl = "https://portfoliohedge.finideas.com/PortFolioPayout/PortfolioService.asmx?WSDL"
    client = Client(wsdl=wsdl, transport=transport)

    result = client.service.Get_EQSymbol()

    try:
        data = json.loads(result)
    except Exception as e:
        print("⚠️ Failed to parse JSON:", e)
        print("Raw result:", result)
        return []

    # Extract only SYMBOL values (if available)
    if isinstance(data, list):
        symbols = [item["SYMBOL"] for item in data if "SYMBOL" in item]
        return symbols
    else:
        return []
    
    
# -------------------------------
# Mutual Fund Beta Calculation Functions
schemes_cache = None
benchmark_data = None
    
def get_all_schemes():
    """Get all mutual fund schemes with caching"""
    global schemes_cache
    if isinstance(schemes_cache, pd.Series) and not schemes_cache.empty:
        return schemes_cache
        
    try:
        print("📡 Fetching mutual fund schemes...")
        url = "https://api.mfapi.in/mf"
        response = requests.get(url, timeout=30)
        if response.status_code == 200:
            schemes_cache = response.json()
            print(f"✅ Loaded {len(schemes_cache)} mutual fund schemes")
            return schemes_cache
        return []
    except Exception as e:
        print(f"❌ Error fetching schemes: {e}")
        logger.error (f"❌ Error fetching schemes: {e}")
        return []

def find_scheme( scheme_name):
    """Find scheme by name (flexible matching)"""
    schemes = get_all_schemes()
    if not schemes:
        return None, None
        
    # Try exact match first
    for scheme in schemes:
        if scheme_name.lower() in scheme['schemeName'].lower():
            return scheme['schemeCode'], scheme['schemeName']
            
    return None, None

def get_nav_data( scheme_code):
    """Get historical NAV data for a scheme"""
    try:
        url = f"https://api.mfapi.in/mf/{scheme_code}"
        response = requests.get(url, timeout=30)
        if response.status_code == 200:
            data = response.json()
            if 'data' in data and data['data']:
                df = pd.DataFrame(data['data'])
                df['date'] = pd.to_datetime(df['date'], format='%d-%m-%Y')
                df['nav'] = pd.to_numeric(df['nav'], errors='coerce')
                df.set_index('date', inplace=True)
                df = df.sort_index().dropna()
                
                # Filter for last 2 years
                two_years_ago = datetime.now() - timedelta(days=730)
                df = df[df.index >= two_years_ago]
                
                return df['nav']
        return None
    except Exception as e:
        print(f"❌ Error fetching NAV for {scheme_code}: {e}")
        logger.info(f"❌ Error fetching NAV for {scheme_code}: {e}")
        return None

def get_benchmark_data():
    """Get Nifty 50 benchmark data - FIXED VERSION"""
    global benchmark_data
    if isinstance(benchmark_data, pd.Series) and not benchmark_data.empty:
        return benchmark_data
        
    try:
        end_date = datetime.now(timezone.utc)
        start_date = end_date - timedelta(days=365)  # 1 years
        
        print("📈 Downloading Nifty 50 benchmark data...")
        nifty_data = yf.download(
            "^NSEI", 
            start=start_date.date(), 
            end=end_date.date(), 
            auto_adjust=True,
            progress=False
        )
        
        # FIX: Ensure we get 1D series
        if isinstance(nifty_data, pd.DataFrame):
            nifty_series = nifty_data['Close']
        else:
            nifty_series = nifty_data
            
        # Convert to 1D series if needed
        if hasattr(nifty_series, 'squeeze'):
            nifty_series = nifty_series.squeeze()
        
        benchmark_data = nifty_series
        print(f"✅ Nifty data: {len(nifty_series)} points")
        return nifty_series
    except Exception as e:
        print(f"❌ Error downloading benchmark: {e}")
        logger.error(f"❌ Error downloading benchmark: {e}")    
        return None

def calculate_beta( nav_series, benchmark_series):
    """Calculate beta between MF and benchmark - FIXED VERSION"""
    try:

        # FIX: Ensure both are 1D series
        if hasattr(nav_series, 'squeeze'):
            nav_series = nav_series.squeeze()
        if hasattr(benchmark_series, 'squeeze'):
            benchmark_series = benchmark_series.squeeze()
        
        # Convert to DataFrames to align dates properly
        mf_df = pd.DataFrame({'nav': nav_series})
        bench_df = pd.DataFrame({'nifty': benchmark_series})
        
        # Merge on date index
        combined = mf_df.merge(bench_df, left_index=True, right_index=True, how='inner')
        
        if len(combined) < 60:
            print(f"   ⚠ Only {len(combined)} common data points")
            return np.nan
        
        # Calculate daily returns
        returns = combined.pct_change().dropna()
        
        if len(returns) < 40:
            return np.nan
        
        # Calculate beta
        covariance = returns['nav'].cov(returns['nifty'])
        variance = returns['nifty'].var()
        
        beta = covariance / variance if variance != 0 else np.nan
        return beta
        
    except Exception as e:
        print(f"❌ Beta calculation error: {e}")
        return np.nan

def calculate_scheme_beta( scheme_name):
    """
    Calculate beta for a single mutual fund
    """
    print(f"\n🔍 Analyzing: {scheme_name}")
    logger.info(f"Analyzing: {scheme_name}")    
    
    # Step 1: Find scheme code
    scheme_code, full_name = find_scheme(scheme_name)
    if not scheme_code:
        print(f"   ❌ Mutual fund not found: {scheme_name}")
        print(f"   💡 Try using the exact name from the fund house")
        return {
            'scheme_name': scheme_name,
            'full_name': 'N/A',
            'scheme_code': 'N/A',
            'beta': 0,
            'data_points': '0',
            'status': 'Failed'
        }
    
    print(f"   ✅ Found: {full_name}")
    
    # Step 2: Get NAV data
    nav_data = get_nav_data(scheme_code)
    if nav_data is None or len(nav_data) < 60:
        print(f"   ❌ Insufficient NAV data: {len(nav_data) if nav_data else 0} points")
        return {
            'scheme_name': scheme_name,
            'full_name': 'N/A',
            'scheme_code': 'N/A',
            'beta': 0,
            'data_points': '0',
            'status': 'Failed'
        }
    
    print(f"   📊 NAV data points: {len(nav_data)}")
    
    # Step 3: Get benchmark data
    benchmark_data = get_benchmark_data()
    if benchmark_data is None or len(benchmark_data) < 60:
        print("   ❌ Insufficient benchmark data")
        return {
            'scheme_name': scheme_name,
            'full_name': 'N/A',
            'scheme_code': 'N/A',
            'beta': 0,
            'data_points': '0',
            'status': 'Failed'
        }
    
    # Step 4: Calculate beta
    beta = calculate_beta(nav_data, benchmark_data)
    
    if not np.isnan(beta):
        print(f"   🎯 Beta: {beta:.4f}")
        return {
            'scheme_name': scheme_name,
            'full_name': full_name,
            'scheme_code': scheme_code,
            'beta': round(beta, 4),
            'data_points': len(nav_data),
            'status': 'Success'
        }
    else:
        print("   ❌ Could not calculate beta")
        logger.info("   ❌ Could not calculate beta")
        return {
            'scheme_name': scheme_name,
            'full_name': full_name,
            'scheme_code': scheme_code,
            'beta': '0',
            'data_points': len(nav_data) if nav_data else 0,
            'status': 'Failed'
        }

def get_beta_for_symbol(symbol, ptype="Stocks"):
    
    # Download NIFTY Index data
    index_series = download_yahoo_adjclose(YAHOO_INDEX_TICKER, START_DATE, END_DATE)
    if index_series is None or index_series.empty:
        st.error("❌ Failed to download index data.")
    else:
        index_series = index_series.dropna().sort_index()

    if ptype == "Stocks":
        print(f"Calculating beta for stock: {symbol}")
        # Use stock beta function
        sym, beta = get_stock_beta(symbol, index_series)
        print(f"Beta for {symbol}: {beta}")
        return sym, beta
    else:
        print(f"💼 Calculating beta for mutual fund scheme code: {symbol}")
        scheme_name = symbol
        # Calculate beta for this scheme
        result = calculate_scheme_beta(scheme_name)
        if result and 'beta' in result:
            beta = result['beta']
            print(f"Beta for {symbol}: {beta}")
            return symbol, beta
            

# -------------------------------
# Page Setup
# -------------------------------
st.set_page_config(page_title="Portfolio Beta & Hedging Calculator", layout="wide", page_icon="📊")

# -------------------------------
# Custom Modern CSS
# -------------------------------
st.markdown("""
<style>
/* ---------- Global ---------- */
html, body, [class*="css"]  {
    font-family: 'Inter', sans-serif;
    background: linear-gradient(145deg, #f7f8fc, #e9ebf1);
    color: #1E1E1E;
}

/* ---------- Header ---------- */
h1 {
    text-align: center;
    font-weight: 800;
    color: #222;
    padding: 1rem 0 0.5rem 0;
    font-size: 2rem;
}

h3 {
    color: #333;
    font-weight: 700;
    margin-top: 1rem;
}



/* ---------- Inputs ---------- */
.stSelectbox, .stNumberInput, .stTextInput, .stFileUploader {
    border-radius: 8px !important;
}

/* ---------- Buttons ---------- */
.stDownloadButton button {
    background-color: #2563eb;
    color: white;
    border-radius: 0.6rem;
    font-weight: 600;
}
.stDownloadButton button:hover {
    background-color: #1d4ed8;
}

/* ---------- DataFrame ---------- */
[data-testid="stDataFrame"] {
    border-radius: 10px;
    overflow: hidden;
}
hr { margin: 0px !important; }
/* ---------- Responsive ---------- */
@media (max-width: 768px) {
    h1 {
        font-size: 1.5rem !important;
    }
    .card {
        padding: 1rem;
    }
}
</style>
""", unsafe_allow_html=True)

# -------------------------------
# App Title
# -------------------------------
st.markdown("<h1>📊 Portfolio Beta & Hedging Calculator</h1>", unsafe_allow_html=True)

# -------------------------------
# Load Data Sources
# -------------------------------
nse_symbols = Get_EQSymbol()

# Load Mutual Fund list from API
scheme_url = "https://api.mfapi.in/mf"
try:
    scheme_resp = requests.get(scheme_url)
    scheme_resp.raise_for_status()
    df_schemes = pd.DataFrame(scheme_resp.json())
    mf_symbols = df_schemes['schemeCode'].tolist()
    mf_name_map = dict(zip(df_schemes['schemeCode'], df_schemes['schemeName']))
except Exception:
    mf_symbols = []
    mf_name_map = {}

# -------------------------------
# Portfolio Type Selection
# -------------------------------

portfolio_type = st.radio("💼 Select Portfolio Type:", ["Stocks", "Mutual Funds"], horizontal=True)
   

# -------------------------------
# Input Method Tabs
# -------------------------------
st.markdown("### 1️⃣ Portfolio Input Method")
portfolio_data = None
active_tab = None
# tab_manual, tab_upload = st.tabs(["✍️ Manual Entry", "📤 Upload File"])
selected_tab = st.radio(
    "Select Input Mode",
    ["✍️ Manual Entry", "📤 Upload File"],
    horizontal=True,
    key="portfolio_tab"
)

# -------------------------------
# Manual Entry
# -------------------------------
if selected_tab == "✍️ Manual Entry":
    active_tab = "manual"
    st.markdown(f"#### Enter {portfolio_type} Manually")
    num_items = st.number_input(f"Number of {portfolio_type}:", min_value=1, max_value=20, value=3)
    items = []

    for i in range(num_items):
        col1, col2 = st.columns(2)
        with col1:
            if portfolio_type == "Stocks":
                symbol = st.selectbox(
                    f"Select Stock Symbol {i+1}",
                    options=nse_symbols,
                    index=nse_symbols.index("RELIANCE") if "RELIANCE" in nse_symbols else 0,
                    key=f"sym_{i}"
                )
                display_name = symbol
            else:
                symbol = st.selectbox(
                    f"Select MF Scheme {i+1}",
                    options=mf_symbols,
                    format_func=lambda code: mf_name_map.get(code, str(code)),
                    key=f"mf_sym_{i}"
                )
                display_name = mf_name_map.get(symbol, str(symbol))
        with col2:
            amount = st.number_input(f"Investment Amount (₹) {i+1}", min_value=0, value=10000, step=1000, key=f"amt_{i}")
        items.append({"SYMBOL": display_name, "AMOUNT": amount})

    portfolio_data = pd.DataFrame(items)
    portfolio_data["TYPE"] = portfolio_type
    st.markdown("#### Your Portfolio:")
    st.dataframe(portfolio_data, use_container_width=True)

# -------------------------------
# CSV/XLSX Upload
# -------------------------------
if selected_tab == "📤 Upload File":
    active_tab = "upload"
    st.markdown(f"#### Upload {portfolio_type} Portfolio File")
    # Sample File
    if portfolio_type == "Stocks":
        st.info("Your file should have columns: SYMBOL, AMOUNT")
        sample_data = pd.DataFrame({"SYMBOL": ["RELIANCE", "INFY"], "AMOUNT": [10000, 15000]})
        file_name = "sample_stock_portfolio.csv"
    else:
        st.info("Your file should have columns: SCHEME_NAME, AMOUNT")
        sample_data = pd.DataFrame({"SCHEME_NAME": [mf_symbols[0] if mf_symbols else ""], "AMOUNT": [15000]})
        file_name = "sample_mf_portfolio.csv"

    csv_buffer = io.StringIO()
    sample_data.to_csv(csv_buffer, index=False)
    st.download_button("📥 Download Sample File", csv_buffer.getvalue(), file_name=file_name, mime="text/csv")

    uploaded_file = st.file_uploader("Upload Portfolio CSV/XLSX", type=['csv', 'xlsx'])
    if uploaded_file is not None:
        if uploaded_file.name.endswith('.csv'):
            portfolio_data = pd.read_csv(uploaded_file)
        else:
            portfolio_data = pd.read_excel(uploaded_file)

        if "SCHEME_NAME" in portfolio_data.columns and "SYMBOL" not in portfolio_data.columns:
            portfolio_data.rename(columns={"SCHEME_NAME": "SYMBOL"}, inplace=True)

        portfolio_data["TYPE"] = portfolio_type
        valid_rows, not_found = [], []

        if portfolio_type == "Mutual Funds":
            for _, row in portfolio_data.iterrows():
                scheme_name = str(row["SYMBOL"]).strip()
                code, name = find_scheme(scheme_name)
                if code:
                    valid_rows.append(row)
                else:
                    not_found.append(scheme_name)
        else:
            for _, row in portfolio_data.iterrows():
                sym = str(row["SYMBOL"]).upper()
                if sym in nse_symbols:
                    valid_rows.append(row)
                else:
                    not_found.append(sym)

        if valid_rows:
            portfolio_data = pd.DataFrame(valid_rows)
        else:
            portfolio_data = pd.DataFrame(columns=["SYMBOL", "AMOUNT", "TYPE"])

        if not_found:
            st.warning(f"⚠️ The following {portfolio_type} were not found: {', '.join(not_found)}")

        st.markdown("#### Uploaded Portfolio:")
        st.dataframe(portfolio_data, use_container_width=True)
    


# -------------------------------
# Hedge Calculation Section
# -------------------------------

if portfolio_data is not None and portfolio_data.shape[0] > 0:
    st.markdown("<h3>2️⃣ Hedge Settings & Black-Scholes Parameters</h3>", unsafe_allow_html=True)
    st.markdown("<hr>", unsafe_allow_html=True)

    # Hedge Protection Level
    st.markdown("### 🛡️ Hedge Protection Level")
    col1, col2 = st.columns(2)
    with col1:
        hedge_percentage = st.selectbox(
            "How much of your portfolio do you want to hedge?",
            [100, 75, 50, 25],
            index=0,
            help="Select the percentage of your portfolio exposure you want to protect"
        )
        st.write(f"**Selected:** {hedge_percentage}% protection")

    with col2:
        st.info(f"""
        **Hedge Percentage Guide:**
        - **100%** → Full protection (most expensive)
        - **75%** → Balanced protection  
        - **50%** → Moderate protection  
        - **25%** → Basic protection (least expensive)
        """)

    st.markdown("<br>", unsafe_allow_html=True)
    
    # Calculate Button
    if st.button("🚀 Calculate Beta & Hedging", type="primary"):
        if "AMOUNT" not in portfolio_data.columns:
            st.error("❌ Portfolio must have 'AMOUNT' column")
        else:
            try:
                # Convert AMOUNT to numeric
                portfolio_data["AMOUNT"] = pd.to_numeric(portfolio_data["AMOUNT"], errors='coerce')
                if portfolio_data["AMOUNT"].isna().any():
                    st.error("❌ Some AMOUNT values are not valid numbers.")
                else:
                    with st.spinner("📊 Calculating betas and advanced hedging costs..."):
                        total_amount = portfolio_data["AMOUNT"].sum()
                        if total_amount == 0:
                            st.error("❌ Total portfolio amount cannot be zero")
                        else:
                            portfolio_data["WEIGHT"] = portfolio_data["AMOUNT"] / total_amount

                            # Calculate betas
                            betas = []
                            for sym in portfolio_data["SYMBOL"]:
                                symbol, beta = get_beta_for_symbol(sym, ptype=portfolio_type)
                                betas.append((symbol, beta))
                            
                            beta_df = pd.DataFrame(betas, columns=["SYMBOL", "BETA"])
                            merged = pd.merge(portfolio_data, beta_df, on="SYMBOL", how="left")
                            merged["WEIGHTED_BETA"] = merged["WEIGHT"] * merged["BETA"]
                            portfolio_beta = merged["WEIGHTED_BETA"].sum()

                            hedging_data_tables = Get_hedge_data(portfolio_beta, total_amount, hedge_percentage)
                            
                            table1 = hedging_data_tables["Table"]
                            row = table1[0]

                            # Extract hedging data
                            hedging_data = {
                                "monthly_expiry": row["Curr_Expiry"].split("T")[0],
                                "quarterly_expiry": row["Qut_Expiry"].split("T")[0],
                                "annual_expiry": row["Annual_Expiry"].split("T")[0],
                                "monthly_put_strike": row["Monthly_Strike"],
                                "quarterly_put_strike": row["Quarterly_Strike"],
                                "annual_put_strike": row["Annual_Strike"],
                                "monthly_cost": row["M_totHedgeCost"],
                                "quarterly_cost": row["Q_totHedgeCost"],
                                "annual_cost": row["A_totHedgeCost"],
                                "monthly_annualized_cost": row["M_costPer"],
                                "quarterly_annualized_cost": row["Q_costPer"],
                                "annual_annualized_cost": row["A_costPer"],
                                "monthly_lots": row["MLot"],
                                "quarterly_lots": row["Qlot"],
                                "annual_lots": row["ALot"],
                                "monthly_premium": row["M_Premium"],
                                "quarterly_premium": row["Q_Premium"],
                                "annual_premium": row["A_Premium"]
                            }

                            # --- Display Results ---
                            st.markdown("<h3>3️⃣ Advanced Hedging Results</h3>", unsafe_allow_html=True)
                            st.markdown("<hr>", unsafe_allow_html=True)
                            st.success("✅ Calculation Complete!")

                            # Protection Level Info
                            st.markdown("### 🧮 Protection Details")
                            col1, col2, col3, col4 = st.columns(4)
                            col1.metric("Total Portfolio Value", f"₹{total_amount:,.2f}")
                            col2.metric("Hedge Percentage", f"{hedge_percentage}%")
                            col3.metric("Portfolio Beta", f"{portfolio_beta:.4f}")
                            col4.metric("Hedge Exposure", f"₹{total_amount * portfolio_beta * (hedge_percentage/100):,.2f}")

                            # Hedging cost details
                            st.markdown("### 💰 Hedging Costs Overview")
                            col1, col2, col3 = st.columns(3)

                            with col1:
                                st.markdown("#### 🗓️ Monthly")
                                st.write(f"**Put Strike:** ₹{hedging_data['monthly_put_strike']:,}")
                                st.write(f"**Expiry:** {hedging_data['monthly_expiry']}")
                                st.write(f"**Cost:** ₹{hedging_data['monthly_cost']:,.2f}")
                                st.write(f"**Annualized Cost:** {hedging_data['monthly_annualized_cost']:.2f}%")
                                st.write(f"**Lots Required:** {hedging_data['monthly_lots']}")
                                st.write(f"**Premium per Lot:** ₹{hedging_data['monthly_premium']}")

                            with col2:
                                st.markdown("#### 📅 Quarterly")
                                st.write(f"**Put Strike:** ₹{hedging_data['quarterly_put_strike']:,}")
                                st.write(f"**Expiry:** {hedging_data['quarterly_expiry']}")
                                st.write(f"**Cost:** ₹{hedging_data['quarterly_cost']:,.2f}")
                                st.write(f"**Annualized Cost:** {hedging_data['quarterly_annualized_cost']:.2f}%")
                                st.write(f"**Lots Required:** {hedging_data['quarterly_lots']}")
                                st.write(f"**Premium per Lot:** ₹{hedging_data['quarterly_premium']}")

                            with col3:
                                st.markdown("#### 📆 Annual")
                                st.write(f"**Put Strike:** ₹{hedging_data['annual_put_strike']:,}")
                                st.write(f"**Expiry:** {hedging_data['annual_expiry']}")
                                st.write(f"**Cost:** ₹{hedging_data['annual_cost']:,.2f}")
                                st.write(f"**Annualized Cost:** {hedging_data['annual_annualized_cost']:.2f}%")
                                st.write(f"**Lots Required:** {hedging_data['annual_lots']}")
                                st.write(f"**Premium per Lot:** ₹{hedging_data['annual_premium']}")

                            # Portfolio Breakdown
                            st.markdown("### 📊 Portfolio Breakdown")
                            st.dataframe(merged)

                            # Scenario Analysis
                            table2 = hedging_data_tables["Table1"]
                            if table2 and len(table2) > 0:
                                st.markdown("### 🎯 Scenario Analysis")
                                scenario_df = pd.DataFrame(table2)
                                scenario_df.columns = [col.lower() for col in scenario_df.columns]
                                for period in ['Monthly', 'Quarterly', 'Annual']:
                                    period_data = scenario_df[scenario_df['period'] == period]
                                    if not period_data.empty:
                                        st.write(f"**{period} Hedging Scenarios:**")
                                        st.dataframe(period_data.drop('period', axis=1))

                            # Download Options
                            st.markdown("### 📥 Download Results")
                            csv = portfolio_data.to_csv(index=False)
                            st.download_button("⬇️ Download Portfolio CSV", csv, "portfolio_results.csv", "text/csv")

                            excel_wb = create_excel_export(portfolio_data, hedging_data, portfolio_beta, total_amount, hedge_percentage)
                            buffer = io.BytesIO()
                            excel_wb.save(buffer)
                            buffer.seek(0)
                            st.download_button("⬇️ Export Full Excel Report", buffer, "portfolio_hedging.xlsx",
                                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"❌ Error during calculation: {e}")
