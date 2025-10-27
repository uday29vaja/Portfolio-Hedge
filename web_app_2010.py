# web_app.py - Streamlit Web Interface for Portfolio Beta
import streamlit as st
import pandas as pd
import numpy as np
import yfinance as yf
from datetime import datetime, timezone, timedelta
import io
import gspread
from google.oauth2.service_account import Credentials
import time
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font
import json
from dotenv import load_dotenv
# Configuration
END_DATE = datetime.now(timezone.utc).date()
START_DATE = END_DATE - timedelta(days=365)
YAHOO_INDEX_TICKER = "^NSEI"


GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1tmlmkpvQK3X15VueaqrNX4xYSd5gudSzhYgJhfIf_QQ/edit?usp=sharing"
# Load environment variables from .env
load_dotenv()

def setup_google_sheets():
    """Connect to Google Sheets"""
    try:
        # Load credentials JSON from environment variable
        cred_json = os.getenv("GOOGLE_CREDS")
        if not cred_json:
            raise ValueError("GOOGLE_CREDS environment variable not found!")
        
        credentials_dict = json.loads(cred_json)

        # Setup Google Sheets API
        scope = ['https://spreadsheets.google.com/feeds',
                 'https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_info(credentials_dict, scopes=scope)
        client = gspread.authorize(creds)
        return client
    except Exception as e:
        st.error(f"Google Sheets connection failed: {e}")
        return None

def percentage_to_float(percent_str):
    """Convert percentage string to float"""
    if not percent_str or percent_str == "N/A":
        return 0.0
    try:
        # Remove % sign and convert to float, then divide by 100
        return float(percent_str.replace('%', '').strip()) / 100
    except:
        return 0.0

def get_hedging_data_from_sheet(portfolio_beta, total_value, hedge_percentage, client):
    """Send beta & portfolio value to sheet, get all calculated data"""
    try:
        spreadsheet = client.open_by_url(GOOGLE_SHEET_URL)
        hedge_sheet = spreadsheet.worksheet("Hedge Cost")
        scenario_sheet = spreadsheet.worksheet("Scenario Analysis")
        
        # INPUT: Send data to Hedge Cost sheet - Convert to native Python types
        hedge_sheet.update('H4', [[float(portfolio_beta)]])     # Portfolio beta
        hedge_sheet.update('H3', [[int(total_value)]])          # Portfolio value
        hedge_sheet.update('H24', [[float(hedge_percentage/100)]])  # Hedge percentage
        
        # Wait for Google Sheet calculations
        st.info("⏳ Running advanced calculations in Google Sheets...")
        time.sleep(8)
        
        # OUTPUT: Read all calculated results from Hedge Cost sheet
        
        # Get common data
        # Get put strikes for each period separately
        monthly_put_strike = hedge_sheet.acell('H14').value or "24,000"
        quarterly_put_strike = hedge_sheet.acell('I14').value or "24,000" 
        annual_put_strike = hedge_sheet.acell('J14').value or "24,000"
        monthly_expiry = hedge_sheet.acell('H16').value or "N/A"
        quarterly_expiry = hedge_sheet.acell('I16').value or "N/A"
        annual_expiry = hedge_sheet.acell('J16').value or "N/A"
        
        # Monthly Hedging
        monthly_cost = hedge_sheet.acell('H28').value or "0"
        monthly_annualized_cost = hedge_sheet.acell('H31').value or "0%"
        monthly_lots = hedge_sheet.acell('H27').value or "N/A"
        monthly_premium = hedge_sheet.acell('H18').value or "N/A"
        
        # Quarterly Hedging  
        quarterly_cost = hedge_sheet.acell('I28').value or "0"
        quarterly_annualized_cost = hedge_sheet.acell('I31').value or "0%"
        quarterly_lots = hedge_sheet.acell('I27').value or "N/A"
        quarterly_premium = hedge_sheet.acell('I18').value or "N/A"
        
        # Annual Hedging
        annual_cost = hedge_sheet.acell('J28').value or "0"
        annual_annualized_cost = hedge_sheet.acell('J31').value or "0%"
        annual_lots = hedge_sheet.acell('J27').value or "N/A"
        annual_premium = hedge_sheet.acell('J18').value or "N/A"
        
        # Get scenario analysis data
        try:
            all_scenarios = []
            
            # Monthly Scenario Analysis - Adjusted ranges
            try:
                monthly_range = scenario_sheet.get('A5:F10')
                for row in monthly_range[1:]:
                    if len(row) >= 6 and row[0]:  # Check if row has data
                        all_scenarios.append({
                            'period': 'Monthly',
                            'scenario': row[0],
                            'end_spot': row[1],
                            'portfolio_wo_hedge': row[2],
                            'put_payoff': row[3],
                            'net_value_hedge': row[4],
                            'hedge_benefit': row[5] if len(row) > 5 else "N/A"
                        })
            except:
                pass
            
            # Quarterly Scenario Analysis - Adjusted ranges
            try:
                quarterly_range = scenario_sheet.get('A17:F22')
                for row in quarterly_range[1:]:
                    if len(row) >= 6 and row[0]:
                        all_scenarios.append({
                            'period': 'Quarterly', 
                            'scenario': row[0],
                            'end_spot': row[1],
                            'portfolio_wo_hedge': row[2],
                            'put_payoff': row[3],
                            'net_value_hedge': row[4],
                            'hedge_benefit': row[5] if len(row) > 5 else "N/A"
                        })
            except:
                pass
            
            # Annual Scenario Analysis - Adjusted ranges
            try:
                annual_range = scenario_sheet.get('A29:F34')
                for row in annual_range[1:]:
                    if len(row) >= 6 and row[0]:
                        all_scenarios.append({
                            'period': 'Annual',
                            'scenario': row[0],
                            'end_spot': row[1],
                            'portfolio_wo_hedge': row[2],
                            'put_payoff': row[3],
                            'net_value_hedge': row[4],
                            'hedge_benefit': row[5] if len(row) > 5 else "N/A"
                        })
            except:
                pass
                    
        except Exception as e:
            st.warning(f"Could not load scenario analysis: {e}")
            all_scenarios = []
        
        return {
            # Common data - now separate put strikes
            'l̥': monthly_put_strike,
            'quarterly_put_strike': quarterly_put_strike, 
            'annual_put_strike': annual_put_strike,
            
        
            'monthly_expiry': monthly_expiry,
            'quarterly_expiry': quarterly_expiry,
            'annual_expiry': annual_expiry,
            
            # Monthly data
            'monthly_cost': float(monthly_cost.replace(',', '')) if monthly_cost and monthly_cost != "N/A" else 0,
            'monthly_annualized_cost': percentage_to_float(monthly_annualized_cost) * 100,  # Convert to percentage
            'monthly_lots': monthly_lots,
            'monthly_premium': monthly_premium,
            
            # Quarterly data
            'quarterly_cost': float(quarterly_cost.replace(',', '')) if quarterly_cost and quarterly_cost != "N/A" else 0,
            'quarterly_annualized_cost': percentage_to_float(quarterly_annualized_cost) * 100,
            'quarterly_lots': quarterly_lots,
            'quarterly_premium': quarterly_premium,
            
            # Annual data
            'annual_cost': float(annual_cost.replace(',', '')) if annual_cost and annual_cost != "N/A" else 0,
            'annual_annualized_cost': percentage_to_float(annual_annualized_cost) * 100,
            'annual_lots': annual_lots,
            'annual_premium': annual_premium,
            
            'scenario_analysis': all_scenarios
        }
        
    except Exception as e:
        st.error(f"Error with Google Sheets: {e}")
        # Fallback calculation
        base_cost = total_value * portfolio_beta * (hedge_percentage/100) * 0.005
        return {
            'put_strike': "24,000",
            'monthly_expiry': "N/A",
            'quarterly_expiry': "N/A", 
            'annual_expiry': "N/A",
            
            'monthly_cost': base_cost / 12,
            'monthly_annualized_cost': (base_cost / 12) / total_value * 100 * 12,
            'monthly_lots': "N/A",
            'monthly_premium': "N/A",
            
            'quarterly_cost': base_cost / 4,
            'quarterly_annualized_cost': (base_cost / 4) / total_value * 100 * 4,
            'quarterly_lots': "N/A",
            'quarterly_premium': "N/A",
            
            'annual_cost': base_cost,
            'annual_annualized_cost': base_cost / total_value * 100,
            'annual_lots': "N/A",
            'annual_premium': "N/A",
            
            'scenario_analysis': []
        }

def download_yahoo_adjclose(ticker, start, end):
    try:
        data = yf.download(
            ticker,
            start=start.isoformat(),
            end=(end + timedelta(days=1)).isoformat(),
            progress=False,
            threads=False
        )
        if data is None or data.empty:
            return None
        return data.get("Adj Close") or data.get("Close")
    except Exception as e:
        st.warning(f"Failed to fetch {ticker}: {e}")
        return None

def compute_beta(stock_series, index_series):
    df = pd.concat([stock_series, index_series], axis=1, join="inner").dropna()
    df = df.sort_index()
    if df.shape[0] < 30:
        return np.nan
    returns = df.pct_change().dropna()
    if returns.shape[0] < 20:
        return np.nan
    cov = returns.cov().iloc[0,1]
    var_index = returns.iloc[:,1].var()
    return cov/var_index if var_index != 0 else np.nan

def get_stock_beta(symbol, index_series):
    yf_ticker = f"{symbol}.NS"
    series = download_yahoo_adjclose(yf_ticker, START_DATE, END_DATE)
    if series is None or series.empty:
        return symbol, np.nan
    beta = compute_beta(series, index_series)
    return symbol, beta

def create_excel_export(portfolio_data, hedging_data, portfolio_beta, total_amount, hedge_percentage):
    """Create Excel file with same display as web app"""
    wb = Workbook()
    
    # Portfolio Summary Sheet
    ws1 = wb.active
    ws1.title = "Portfolio Summary"
    
    # Header
    ws1['A1'] = "Portfolio Beta & Hedging Calculator - Results"
    ws1['A1'].font = Font(bold=True, size=14)
    
    # Protection Details
    ws1['A3'] = "Protection Details"
    ws1['A3'].font = Font(bold=True)
    ws1['A4'] = "Total Portfolio Value"
    ws1['B4'] = f"₹{total_amount:,.2f}"
    ws1['A5'] = "Hedge Percentage"
    ws1['B5'] = f"{hedge_percentage}%"
    ws1['A6'] = "Portfolio Beta"
    ws1['B6'] = f"{portfolio_beta:.4f}"
    ws1['A7'] = "Hedge Exposure"
    ws1['B7'] = f"₹{total_amount * portfolio_beta * (hedge_percentage/100):,.2f}"
    
    # Portfolio Breakdown
    ws1['A9'] = "Portfolio Breakdown"
    ws1['A9'].font = Font(bold=True)
    
    # Headers
    headers = list(portfolio_data.columns)
    for col, header in enumerate(headers, 1):
        ws1.cell(row=10, column=col, value=header)
        ws1.cell(row=10, column=col).font = Font(bold=True)
    
    # Data
    for row, (_, data) in enumerate(portfolio_data.iterrows(), 11):
        for col, value in enumerate(data, 1):
            ws1.cell(row=row, column=col, value=value)
    
    # Hedging Costs Sheet
    ws2 = wb.create_sheet("Hedging Costs")
    
    ws2['A1'] = "Hedging Costs & Details"
    ws2['A1'].font = Font(bold=True, size=14)
    
    # Headers for hedging costs
    hedging_headers = ["Option Type", "Put Strike", "Expiry", "Cost", "Annualized Cost %", "Lots Required", "Premium per Lot"]
    for col, header in enumerate(hedging_headers, 1):
        ws2.cell(row=3, column=col, value=header)
        ws2.cell(row=3, column=col).font = Font(bold=True)
    
    # Monthly data
    ws2.cell(row=4, column=1, value="Monthly")
    ws2.cell(row=4, column=2, value=f"₹{hedging_data['monthly_put_strike']}")
    ws2.cell(row=4, column=3, value=hedging_data['monthly_expiry'])
    ws2.cell(row=4, column=4, value=f"₹{hedging_data['monthly_cost']:,.2f}")
    ws2.cell(row=4, column=5, value=f"{hedging_data['monthly_annualized_cost']:.2f}%")
    ws2.cell(row=4, column=6, value=hedging_data['monthly_lots'])
    ws2.cell(row=4, column=7, value=f"₹{hedging_data['monthly_premium']}")
    
    # Quarterly data
    ws2.cell(row=5, column=1, value="Quarterly")
    ws2.cell(row=5, column=2, value=f"₹{hedging_data['quarterly_put_strike']}")
    ws2.cell(row=5, column=3, value=hedging_data['quarterly_expiry'])
    ws2.cell(row=5, column=4, value=f"₹{hedging_data['quarterly_cost']:,.2f}")
    ws2.cell(row=5, column=5, value=f"{hedging_data['quarterly_annualized_cost']:.2f}%")
    ws2.cell(row=5, column=6, value=hedging_data['quarterly_lots'])
    ws2.cell(row=5, column=7, value=f"₹{hedging_data['quarterly_premium']}")
    
    # Annual data
    ws2.cell(row=6, column=1, value="Annual")
    ws2.cell(row=6, column=2, value=f"₹{hedging_data['annual_put_strike']}")
    ws2.cell(row=6, column=3, value=hedging_data['annual_expiry'])
    ws2.cell(row=6, column=4, value=f"₹{hedging_data['annual_cost']:,.2f}")
    ws2.cell(row=6, column=5, value=f"{hedging_data['annual_annualized_cost']:.2f}%")
    ws2.cell(row=6, column=6, value=hedging_data['annual_lots'])
    ws2.cell(row=6, column=7, value=f"₹{hedging_data['annual_premium']}")
    
    # Auto-adjust column widths
    for ws in [ws1, ws2]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

# Streamlit Web Interface
st.set_page_config(page_title="Portfolio Beta Calculator", layout="wide")
st.title("📊 Advanced Portfolio Beta & Hedging Calculator")
st.write("Calculate portfolio beta and advanced hedging costs using Black-Scholes model")

# Input Section
st.header("1. Portfolio Input")

input_method = st.radio("Choose input method:", ["Manual Entry", "CSV Upload"])

portfolio_data = None

if input_method == "Manual Entry":
    st.subheader("Enter Stocks Manually")
    
    num_stocks = st.number_input("Number of stocks:", min_value=1, max_value=20, value=3)
    
    stocks = []
    for i in range(num_stocks):
        col1, col2 = st.columns(2)
        with col1:
            symbol = st.text_input(f"Stock Symbol {i+1}", value="RELIANCE", key=f"sym_{i}")
        with col2:
            amount = st.number_input(f"Investment Amount (₹) {i+1}", min_value=0, value=10000, key=f"amt_{i}")
        stocks.append({"SYMBOL": symbol.upper().replace('.NS', ''), "AMOUNT": amount})
    
    if stocks:
        portfolio_data = pd.DataFrame(stocks)
        st.write("Your Portfolio:")
        st.dataframe(portfolio_data)

else:  # CSV Upload
    st.subheader("Upload Portfolio CSV")
    st.info("Your CSV should have columns: SYMBOL, AMOUNT")
    
    uploaded_file = st.file_uploader("Choose CSV file", type=['csv'])
    
    if uploaded_file is not None:
        portfolio_data = pd.read_csv(uploaded_file)
        st.write("Uploaded Portfolio:")
        st.dataframe(portfolio_data)
# Equity Symbols Reference Section - CORRECTED VERSION
st.header("📋 Equity Symbols Reference")

try:
    # Check if file exists and load it
    if os.path.exists("EQUITY_L.csv"):
        scrips_df = pd.read_csv("EQUITY_L.csv")
        total_symbols = len(scrips_df)
        
        st.success(f"**{total_symbols} equity symbols available**")
        st.write("Reference list of all available trading symbols")
        
        # Download option
        with open("EQUITY_L.csv", "rb") as file:
            file_bytes = file.read()
        
        st.download_button(
            label="📥 Download Equity Symbols (CSV)",
            data=file_bytes,
            file_name="EQUITY_Symbols.csv",
            mime="text/csv"
        )
    else:
        st.info("Equity symbols file not found - download available when file is present")

except Exception as e:
    st.info("Equity reference data will be available when the symbols file is present")

# Calculation section with Google Sheets
if portfolio_data is not None:
    # ... rest of your existing code continues here
    st.header("2. Calculate Beta & Advanced Hedging")
    
    # Add Hedge Percentage Selection
    st.subheader("🛡️ Hedge Protection Level")
    col1, col2 = st.columns(2)
    with col1:
        hedge_percentage = st.selectbox(
            "How much portfolio do you want to hedge?",
            [100, 75, 50, 25],
            index=0,
            help="Select the percentage of your portfolio exposure you want to protect"
        )
        st.write(f"**Selected: {hedge_percentage}% protection**")
    
    with col2:
        st.info(f"""
        **Hedge Percentage Guide:**
        - **100%**: Full protection (most expensive)
        - **75%**: Balanced protection
        - **50%**: Moderate protection  
        - **25%**: Basic protection (least expensive)
        """)
    
    if st.button("🚀 Calculate Portfolio Beta & Hedging", type="primary"):
        if "AMOUNT" not in portfolio_data.columns:
            st.error("❌ Portfolio must have 'AMOUNT' column")
        else:
            try:
                # Convert AMOUNT to numbers
                portfolio_data["AMOUNT"] = pd.to_numeric(portfolio_data["AMOUNT"], errors='coerce')
                
                if portfolio_data["AMOUNT"].isna().any():
                    st.error("❌ Some AMOUNT values are not valid numbers.")
                else:
                    with st.spinner("📊 Calculating betas and advanced hedging costs..."):
                        # Calculate weights and beta
                        total_amount = portfolio_data["AMOUNT"].sum()
                        if total_amount == 0:
                            st.error("❌ Total portfolio amount cannot be zero")
                        else:
                            portfolio_data["WEIGHT"] = portfolio_data["AMOUNT"] / total_amount

                            # Download index data and calculate beta
                            index_series = download_yahoo_adjclose(YAHOO_INDEX_TICKER, START_DATE, END_DATE)
                            
                            if index_series is None or index_series.empty:
                                st.error("❌ Failed to download index data.")
                            else:
                                index_series = index_series.dropna().sort_index()

                                # Calculate betas
                                betas = []
                                for sym in portfolio_data["SYMBOL"]:
                                    symbol, beta = get_stock_beta(sym, index_series)
                                    betas.append((symbol, beta))
                                
                                # Merge results
                                beta_df = pd.DataFrame(betas, columns=["SYMBOL", "BETA"])
                                merged = pd.merge(portfolio_data, beta_df, on="SYMBOL", how="left")
                                merged["WEIGHTED_BETA"] = merged["WEIGHT"] * merged["BETA"]
                                portfolio_beta = merged["WEIGHTED_BETA"].sum()

                                # Connect to Google Sheets for advanced calculations
                                st.info("🔗 Connecting to Google Sheets for Black-Scholes calculations...")
                                client = setup_google_sheets()
                                
                                if client:
                                    hedging_data = get_hedging_data_from_sheet(portfolio_beta, total_amount, hedge_percentage, client)
                                    
                                    # Display Results
                                    st.header("3. Advanced Hedging Results")
                                    st.success("✅ Calculation Complete!")
                                    
                                    # Protection Level Info
                                    st.subheader("🛡️ Protection Details")
                                    col1, col2, col3, col4 = st.columns(4)
                                    with col1:
                                        st.metric("Total Portfolio Value", f"₹{total_amount:,.2f}")
                                    with col2:
                                        st.metric("Hedge Percentage", f"{hedge_percentage}%")
                                    with col3:
                                        st.metric("Portfolio Beta", f"{portfolio_beta:.4f}")
                                    with col4:
                                        st.metric("Hedge Exposure", f"₹{total_amount * portfolio_beta * (hedge_percentage/100):,.2f}")
                                    
                                    # Hedging Costs Breakdown
                                    st.subheader("💰 Hedging Costs & Details")
                                    
                                    col1, col2, col3 = st.columns(3)
                                    with col1:
                                        # For Monthly section:
                                        st.metric("Put Strike", f"₹{hedging_data['monthly_put_strike']}")
                                        st.metric("Expiry", hedging_data['monthly_expiry'])
                                        st.metric("Cost", f"₹{hedging_data['monthly_cost']:,.2f}")
                                        st.metric("Annualized Cost %", f"{hedging_data['monthly_annualized_cost']:.2f}%")
                                        st.metric("Lots Required", hedging_data['monthly_lots'])
                                        st.metric("Premium per Lot", f"₹{hedging_data['monthly_premium']}")
                                    
                                    with col2:
                                        st.write("**📊 Quarterly Protection**")
                                        st.metric("Put Strike", f"₹{hedging_data['quarterly_put_strike']}")
                                        st.metric("Expiry", hedging_data['quarterly_expiry'])
                                        st.metric("Cost", f"₹{hedging_data['quarterly_cost']:,.2f}")
                                        st.metric("Annualized Cost %", f"{hedging_data['quarterly_annualized_cost']:.2f}%")
                                        st.metric("Lots Required", hedging_data['quarterly_lots'])
                                        st.metric("Premium per Lot", f"₹{hedging_data['quarterly_premium']}")
                                    
                                    with col3:
                                        st.write("**📈 Annual Protection**")
                                        st.metric("Put Strike", f"₹{hedging_data['annual_put_strike']}")
                                        st.metric("Expiry", hedging_data['annual_expiry'])
                                        st.metric("Cost", f"₹{hedging_data['annual_cost']:,.2f}")
                                        st.metric("Annualized Cost %", f"{hedging_data['annual_annualized_cost']:.2f}%")
                                        st.metric("Lots Required", hedging_data['annual_lots'])
                                        st.metric("Premium per Lot", f"₹{hedging_data['annual_premium']}")
                                    
                                    # Portfolio Breakdown
                                    st.subheader("📈 Portfolio Breakdown")
                                    st.dataframe(merged)
                                    
                                    # Scenario Analysis
                                    if hedging_data['scenario_analysis']:
                                        st.subheader("🎯 Scenario Analysis")
                                        scenario_df = pd.DataFrame(hedging_data['scenario_analysis'])
                                        
                                        # Display by period
                                        for period in ['Monthly', 'Quarterly', 'Annual']:
                                            period_data = scenario_df[scenario_df['period'] == period]
                                            if not period_data.empty:
                                                st.write(f"**{period} Hedging Scenarios:**")
                                                display_data = period_data.drop('period', axis=1)
                                                st.dataframe(display_data, use_container_width=True)
                                    else:
                                        st.info("📊 Scenario analysis data will be available when Google Sheets calculations are complete")
                                    
                                    # Download Results
                                    st.subheader("📥 Download Results")

                                    col1, col2 = st.columns(2)

                                    with col1:
                                        # Original CSV download
                                        csv = merged.to_csv(index=False)
                                        st.download_button(
                                            label="Download Portfolio Results as CSV",
                                            data=csv,
                                            file_name="portfolio_beta_results.csv",
                                            mime="text/csv"
                                        )

                                    with col2:
                                        # New Excel Export
                                        excel_wb = create_excel_export(merged, hedging_data, portfolio_beta, total_amount, hedge_percentage)
                                        
                                        # Save to bytes
                                        excel_buffer = io.BytesIO()
                                        excel_wb.save(excel_buffer)
                                        excel_buffer.seek(0)
                                        
                                        st.download_button(
                                            label="📊 Export Full Report to Excel",
                                            data=excel_buffer,
                                            file_name="portfolio_hedging_full_report.xlsx",
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                        )
                                else:
                                    st.error("❌ Could not connect to Google Sheets. Using fallback calculation.")
                                    # Fallback display
                                    base_cost = total_amount * portfolio_beta * (hedge_percentage/100) * 0.01
                                    col1, col2 = st.columns(2)
                                    with col1:
                                        st.metric("Total Portfolio Value", f"₹{total_amount:,.2f}")
                                        st.metric("Weighted Average Beta", f"{portfolio_beta:.4f}")
                                        st.metric("Hedge Percentage", f"{hedge_percentage}%")
                                    with col2:
                                        st.metric("Estimated Monthly Cost", f"₹{base_cost/12:,.2f}")
                                        st.metric("Estimated Quarterly Cost", f"₹{base_cost/4:,.2f}")
                                        st.metric("Estimated Annual Cost", f"₹{base_cost:,.2f}")
            
            except Exception as e:
                st.error(f"❌ Error in calculation: {str(e)}")